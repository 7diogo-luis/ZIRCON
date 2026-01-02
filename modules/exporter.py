"""Copyright (c) 2025-present Diogo Luís.

Distributed under the MIT software license, see the accompanying
file LICENSE or http://www.opensource.org/licenses/mit-license.php.
"""

from openpyxl import load_workbook
import os
import pickle
import platform


def exporter(IP, mode):
    """Export data contained in the IP dictionary.

    Parameters
    ----------
    IP : dict
        Station's interlocking program.
    mode : str
        "pickle" if the data is to be exported to a pickled file, "xlsx" if the
        data is to be exported to a .xlsx file.
    """
    if mode == 'pickle':
        exportPickle(IP)

    elif mode == 'xlsx':
        exportXlsx(IP)


def exportPickle(IP):
    """Pickle the IP dictionary and save it to a byte stream file.

    Parameters
    ----------
    IP : dict
        Station's interlocking program.
    """
    if platform.system() == 'Windows':
        save_path = os.getcwd() + '\\stations\\output\\'
    else:
        save_path = os.getcwd() + '/stations/output/'

    file_lbl = IP['COVER']['station_lbl'] + '_Interlocking_Program'

    with open((save_path + file_lbl), 'ab') as file:
        pickle.dump(IP, file)


def exportXlsx(IP):
    """Create a copy of the .xlsx template and populate it with IP data.

    Parameters
    ----------
    IP : dict
        Station's interlocking program.
    """
    if platform.system() == 'Windows':
        path = os.getcwd() + '\\templates\\interlocking_program.xlsx'
    else:
        path = os.getcwd() + '/templates/interlocking_program.xlsx'

    workbook = load_workbook(path)

    sta_lbl_str = (IP['COVER']['station_name'] + ' (' +
                   IP['COVER']['station_lbl'] + ')')
    workbook['Cover']['A8'] = sta_lbl_str
    workbook['Cover']['A9'] = IP['COVER']['interlocking_name']
    workbook['Cover']['F3'] = IP['COVER']['sw_version']
    workbook['Cover']['A19'] = IP['COVER']['encoding_author']
    workbook['Cover']['A20'] = IP['COVER']['date']

    row_num = 5

    for movement in IP['CIRCULATION']['normal_entry']:
        row_num += 1
        workbook['Normal Entries (circulation)'].cell(row_num, 1, movement
                                                      ['rt_ID'])
        workbook['Normal Entries (circulation)'].cell(row_num, 2, movement
                                                      ['origin_sig'])
        workbook['Normal Entries (circulation)'].cell(row_num, 3, movement
                                                      ['origin_loc'])
        workbook['Normal Entries (circulation)'].cell(row_num, 4, movement
                                                      ['destination_sig'])
        workbook['Normal Entries (circulation)'].cell(row_num, 5, movement
                                                      ['destination_loc'])
        workbook['Normal Entries (circulation)'].cell(row_num, 6, movement
                                                      ['regime'])
        workbook['Normal Entries (circulation)'].cell(row_num, 7, movement
                                                      ['alt_ol'])
        workbook['Normal Entries (circulation)'].cell(row_num, 8, movement
                                                      ['alt_rt'])
        workbook['Normal Entries (circulation)'].cell(row_num, 9, movement
                                                      ['obs'])
        workbook['Normal Entries (circulation)'].cell(row_num, 10, movement
                                                      ['switches']['rt']
                                                      ['normal'])
        workbook['Normal Entries (circulation)'].cell(row_num, 11, movement
                                                      ['switches']['rt']
                                                      ['reverse'])
        workbook['Normal Entries (circulation)'].cell(row_num, 12, movement
                                                      ['switches']['ol']
                                                      ['normal'])
        workbook['Normal Entries (circulation)'].cell(row_num, 13, movement
                                                      ['switches']['ol']
                                                      ['reverse'])
        workbook['Normal Entries (circulation)'].cell(row_num, 14, movement
                                                      ['switches']['fp']['rt']
                                                      ['normal'])
        workbook['Normal Entries (circulation)'].cell(row_num, 15, movement
                                                      ['switches']['fp']['rt']
                                                      ['reverse'])
        workbook['Normal Entries (circulation)'].cell(row_num, 16, movement
                                                      ['switches']['fp']['ol']
                                                      ['normal'])
        workbook['Normal Entries (circulation)'].cell(row_num, 17, movement
                                                      ['switches']['fp']['ol']
                                                      ['reverse'])
        workbook['Normal Entries (circulation)'].cell(row_num, 18, movement
                                                      ['sections']['rt'])
        workbook['Normal Entries (circulation)'].cell(row_num, 19, movement
                                                      ['sections']['ol'])
        workbook['Normal Entries (circulation)'].cell(row_num, 20, movement
                                                      ['sections']['fp']['rt']
                                                      ['vital'])
        workbook['Normal Entries (circulation)'].cell(row_num, 21, movement
                                                      ['sections']['fp']['rt']
                                                      ['sub_vital'])
        workbook['Normal Entries (circulation)'].cell(row_num, 22, movement
                                                      ['sections']['fp']['rt']
                                                      ['remote'])
        workbook['Normal Entries (circulation)'].cell(row_num, 23, movement
                                                      ['sections']['fp']['ol']
                                                      ['vital'])
        workbook['Normal Entries (circulation)'].cell(row_num, 24, movement
                                                      ['sections']['fp']['ol']
                                                      ['sub_vital'])
        workbook['Normal Entries (circulation)'].cell(row_num, 25, movement
                                                      ['sections']['fp']['ol']
                                                      ['remote'])
        workbook['Normal Entries (circulation)'].cell(row_num, 26, movement
                                                      ['blocks']['up'])
        workbook['Normal Entries (circulation)'].cell(row_num, 27, movement
                                                      ['blocks']['down'])

    row_num = 5

    for movement in IP['CIRCULATION']['normal_exit']:
        row_num += 1
        workbook['Normal Exits (circulation)'].cell(row_num, 1, movement
                                                    ['rt_ID'])
        workbook['Normal Exits (circulation)'].cell(row_num, 2, movement
                                                    ['origin_sig'])
        workbook['Normal Exits (circulation)'].cell(row_num, 3, movement
                                                    ['origin_loc'])
        workbook['Normal Exits (circulation)'].cell(row_num, 4, movement
                                                    ['destination_sig'])
        workbook['Normal Exits (circulation)'].cell(row_num, 5, movement
                                                    ['destination_loc'])
        workbook['Normal Exits (circulation)'].cell(row_num, 6, movement
                                                    ['regime'])
        workbook['Normal Exits (circulation)'].cell(row_num, 7, movement
                                                    ['alt_ol'])
        workbook['Normal Exits (circulation)'].cell(row_num, 8, movement
                                                    ['alt_rt'])
        workbook['Normal Exits (circulation)'].cell(row_num, 9, movement
                                                    ['obs'])
        workbook['Normal Exits (circulation)'].cell(row_num, 10, movement
                                                    ['switches']['rt']
                                                    ['normal'])
        workbook['Normal Exits (circulation)'].cell(row_num, 11, movement
                                                    ['switches']['rt']
                                                    ['reverse'])
        workbook['Normal Exits (circulation)'].cell(row_num, 12, movement
                                                    ['switches']['ol']
                                                    ['normal'])
        workbook['Normal Exits (circulation)'].cell(row_num, 13, movement
                                                    ['switches']['ol']
                                                    ['reverse'])
        workbook['Normal Exits (circulation)'].cell(row_num, 14, movement
                                                    ['switches']['fp']['rt']
                                                    ['normal'])
        workbook['Normal Exits (circulation)'].cell(row_num, 15, movement
                                                    ['switches']['fp']['rt']
                                                    ['reverse'])
        workbook['Normal Exits (circulation)'].cell(row_num, 16, movement
                                                    ['switches']['fp']['ol']
                                                    ['normal'])
        workbook['Normal Exits (circulation)'].cell(row_num, 17, movement
                                                    ['switches']['fp']['ol']
                                                    ['reverse'])
        workbook['Normal Exits (circulation)'].cell(row_num, 18, movement
                                                    ['sections']['rt'])
        workbook['Normal Exits (circulation)'].cell(row_num, 19, movement
                                                    ['sections']['ol'])
        workbook['Normal Exits (circulation)'].cell(row_num, 20, movement
                                                    ['sections']['fp']['rt']
                                                    ['vital'])
        workbook['Normal Exits (circulation)'].cell(row_num, 21, movement
                                                    ['sections']['fp']['rt']
                                                    ['sub_vital'])
        workbook['Normal Exits (circulation)'].cell(row_num, 22, movement
                                                    ['sections']['fp']['rt']
                                                    ['remote'])
        workbook['Normal Exits (circulation)'].cell(row_num, 23, movement
                                                    ['sections']['fp']['ol']
                                                    ['vital'])
        workbook['Normal Exits (circulation)'].cell(row_num, 24, movement
                                                    ['sections']['fp']['ol']
                                                    ['sub_vital'])
        workbook['Normal Exits (circulation)'].cell(row_num, 25, movement
                                                    ['sections']['fp']['ol']
                                                    ['remote'])
        workbook['Normal Exits (circulation)'].cell(row_num, 26, movement
                                                    ['blocks']['up'])
        workbook['Normal Exits (circulation)'].cell(row_num, 27, movement
                                                    ['blocks']['down'])

    row_num = 5

    for movement in IP['CIRCULATION']['reverse_entry']:
        row_num += 1
        workbook['Reverse Entries (circulation)'].cell(row_num, 1, movement
                                                       ['rt_ID'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 2, movement
                                                       ['origin_sig'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 3, movement
                                                       ['origin_loc'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 4, movement
                                                       ['destination_sig'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 5, movement
                                                       ['destination_loc'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 6, movement
                                                       ['regime'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 7, movement
                                                       ['alt_ol'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 8, movement
                                                       ['alt_rt'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 9, movement
                                                       ['obs'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 10, movement
                                                       ['switches']['rt']
                                                       ['normal'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 11, movement
                                                       ['switches']['rt']
                                                       ['reverse'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 12, movement
                                                       ['switches']['ol']
                                                       ['normal'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 13, movement
                                                       ['switches']['ol']
                                                       ['reverse'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 14, movement
                                                       ['switches']['fp']['rt']
                                                       ['normal'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 15, movement
                                                       ['switches']['fp']['rt']
                                                       ['reverse'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 16, movement
                                                       ['switches']['fp']['ol']
                                                       ['normal'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 17, movement
                                                       ['switches']['fp']['ol']
                                                       ['reverse'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 18, movement
                                                       ['sections']['rt'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 19, movement
                                                       ['sections']['ol'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 20, movement
                                                       ['sections']['fp']['rt']
                                                       ['vital'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 21, movement
                                                       ['sections']['fp']['rt']
                                                       ['sub_vital'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 22, movement
                                                       ['sections']['fp']['rt']
                                                       ['remote'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 23, movement
                                                       ['sections']['fp']['ol']
                                                       ['vital'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 24, movement
                                                       ['sections']['fp']['ol']
                                                       ['sub_vital'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 25, movement
                                                       ['sections']['fp']['ol']
                                                       ['remote'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 26, movement
                                                       ['blocks']['up'])
        workbook['Reverse Entries (circulation)'].cell(row_num, 27, movement
                                                       ['blocks']['down'])

    row_num = 5

    for movement in IP['CIRCULATION']['reverse_exit']:
        row_num += 1
        workbook['Reverse Exits (circulation)'].cell(row_num, 1, movement
                                                     ['rt_ID'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 2, movement
                                                     ['origin_sig'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 3, movement
                                                     ['origin_loc'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 4, movement
                                                     ['destination_sig'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 5, movement
                                                     ['destination_loc'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 6, movement
                                                     ['regime'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 7, movement
                                                     ['alt_ol'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 8, movement
                                                     ['alt_rt'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 9, movement
                                                     ['obs'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 10, movement
                                                     ['switches']['rt']
                                                     ['normal'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 11, movement
                                                     ['switches']['rt']
                                                     ['reverse'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 12, movement
                                                     ['switches']['ol']
                                                     ['normal'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 13, movement
                                                     ['switches']['ol']
                                                     ['reverse'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 14, movement
                                                     ['switches']['fp']['rt']
                                                     ['normal'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 15, movement
                                                     ['switches']['fp']['rt']
                                                     ['reverse'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 16, movement
                                                     ['switches']['fp']['ol']
                                                     ['normal'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 17, movement
                                                     ['switches']['fp']['ol']
                                                     ['reverse'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 18, movement
                                                     ['sections']['rt'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 19, movement
                                                     ['sections']['ol'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 20, movement
                                                     ['sections']['fp']['rt']
                                                     ['vital'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 21, movement
                                                     ['sections']['fp']['rt']
                                                     ['sub_vital'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 22, movement
                                                     ['sections']['fp']['rt']
                                                     ['remote'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 23, movement
                                                     ['sections']['fp']['ol']
                                                     ['vital'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 24, movement
                                                     ['sections']['fp']['ol']
                                                     ['sub_vital'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 25, movement
                                                     ['sections']['fp']['ol']
                                                     ['remote'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 26, movement
                                                     ['blocks']['up'])
        workbook['Reverse Exits (circulation)'].cell(row_num, 27, movement
                                                     ['blocks']['down'])

    row_num = 5

    for movement in IP['SHUNT']['forward']:
        row_num += 1
        workbook['Forward (shunt)'].cell(row_num, 1, movement['rt_ID'])
        workbook['Forward (shunt)'].cell(row_num, 2, movement['origin_sig'])
        workbook['Forward (shunt)'].cell(row_num, 3, movement['origin_loc'])
        workbook['Forward (shunt)'].cell(row_num, 4, movement
                                         ['destination_sig'])
        workbook['Forward (shunt)'].cell(row_num, 5, movement
                                         ['destination_loc'])
        workbook['Forward (shunt)'].cell(row_num, 6, movement['regime'])
        workbook['Forward (shunt)'].cell(row_num, 7, movement['alt_ol'])
        workbook['Forward (shunt)'].cell(row_num, 8, movement['alt_rt'])
        workbook['Forward (shunt)'].cell(row_num, 9, movement['obs'])
        workbook['Forward (shunt)'].cell(row_num, 10, movement['switches']
                                         ['rt']['normal'])
        workbook['Forward (shunt)'].cell(row_num, 11, movement['switches']
                                         ['rt']['reverse'])
        workbook['Forward (shunt)'].cell(row_num, 12, movement['switches']
                                         ['ol']['normal'])
        workbook['Forward (shunt)'].cell(row_num, 13, movement['switches']
                                         ['ol']['reverse'])
        workbook['Forward (shunt)'].cell(row_num, 14, movement['switches']
                                         ['fp']['rt']['normal'])
        workbook['Forward (shunt)'].cell(row_num, 15, movement['switches']
                                         ['fp']['rt']['reverse'])
        workbook['Forward (shunt)'].cell(row_num, 16, movement['switches']
                                         ['fp']['ol']['normal'])
        workbook['Forward (shunt)'].cell(row_num, 17, movement['switches']
                                         ['fp']['ol']['reverse'])
        workbook['Forward (shunt)'].cell(row_num, 18, movement['sections']
                                         ['rt'])
        workbook['Forward (shunt)'].cell(row_num, 19, movement['sections']
                                         ['ol'])
        workbook['Forward (shunt)'].cell(row_num, 20, movement['sections']
                                         ['fp']['rt']['vital'])
        workbook['Forward (shunt)'].cell(row_num, 21, movement['sections']
                                         ['fp']['rt']['sub_vital'])
        workbook['Forward (shunt)'].cell(row_num, 22, movement['sections']
                                         ['fp']['rt']['remote'])
        workbook['Forward (shunt)'].cell(row_num, 23, movement['sections']
                                         ['fp']['ol']['vital'])
        workbook['Forward (shunt)'].cell(row_num, 24, movement['sections']
                                         ['fp']['ol']['sub_vital'])
        workbook['Forward (shunt)'].cell(row_num, 25, movement['sections']
                                         ['fp']['ol']['remote'])
        workbook['Forward (shunt)'].cell(row_num, 26, movement['blocks']
                                         ['up'])
        workbook['Forward (shunt)'].cell(row_num, 27, movement['blocks']
                                         ['down'])

    row_num = 5

    for movement in IP['SHUNT']['backward']:
        row_num += 1
        workbook['Backward (shunt)'].cell(row_num, 1, movement['rt_ID'])
        workbook['Backward (shunt)'].cell(row_num, 2, movement['origin_sig'])
        workbook['Backward (shunt)'].cell(row_num, 3, movement['origin_loc'])
        workbook['Backward (shunt)'].cell(row_num, 4, movement
                                          ['destination_sig'])
        workbook['Backward (shunt)'].cell(row_num, 5, movement
                                          ['destination_loc'])
        workbook['Backward (shunt)'].cell(row_num, 6, movement['regime'])
        workbook['Backward (shunt)'].cell(row_num, 7, movement['alt_ol'])
        workbook['Backward (shunt)'].cell(row_num, 8, movement['alt_rt'])
        workbook['Backward (shunt)'].cell(row_num, 9, movement['obs'])
        workbook['Backward (shunt)'].cell(row_num, 10, movement['switches']
                                          ['rt']['normal'])
        workbook['Backward (shunt)'].cell(row_num, 11, movement['switches']
                                          ['rt']['reverse'])
        workbook['Backward (shunt)'].cell(row_num, 12, movement['switches']
                                          ['ol']['normal'])
        workbook['Backward (shunt)'].cell(row_num, 13, movement['switches']
                                          ['ol']['reverse'])
        workbook['Backward (shunt)'].cell(row_num, 14, movement['switches']
                                          ['fp']['rt']['normal'])
        workbook['Backward (shunt)'].cell(row_num, 15, movement['switches']
                                          ['fp']['rt']['reverse'])
        workbook['Backward (shunt)'].cell(row_num, 16, movement['switches']
                                          ['fp']['ol']['normal'])
        workbook['Backward (shunt)'].cell(row_num, 17, movement['switches']
                                          ['fp']['ol']['reverse'])
        workbook['Backward (shunt)'].cell(row_num, 18, movement['sections']
                                          ['rt'])
        workbook['Backward (shunt)'].cell(row_num, 19, movement['sections']
                                          ['ol'])
        workbook['Backward (shunt)'].cell(row_num, 20, movement['sections']
                                          ['fp']['rt']['vital'])
        workbook['Backward (shunt)'].cell(row_num, 21, movement['sections']
                                          ['fp']['rt']['sub_vital'])
        workbook['Backward (shunt)'].cell(row_num, 22, movement['sections']
                                          ['fp']['rt']['remote'])
        workbook['Backward (shunt)'].cell(row_num, 23, movement['sections']
                                          ['fp']['ol']['vital'])
        workbook['Backward (shunt)'].cell(row_num, 24, movement['sections']
                                          ['fp']['ol']['sub_vital'])
        workbook['Backward (shunt)'].cell(row_num, 25, movement['sections']
                                          ['fp']['ol']['remote'])
        workbook['Backward (shunt)'].cell(row_num, 26, movement['blocks']
                                          ['up'])
        workbook['Backward (shunt)'].cell(row_num, 27, movement['blocks']
                                          ['down'])

    row_num = 2

    for ol_delay in IP['DELAYS']['overlap']:
        row_num += 1
        workbook['Delays'].cell(row_num, 1, ol_delay['track'])
        workbook['Delays'].cell(row_num, 2, ol_delay['delay'])

    row_num = 2

    for ARC in IP['DELAYS']['approach_rt_cncl']:
        row_num += 1
        workbook['Delays'].cell(row_num, 3, ARC['signal'])
        workbook['Delays'].cell(row_num, 4, ARC['delay'])

    row_num = 2

    for ERC in IP['DELAYS']['emerg_rt_cncl']:
        row_num += 1
        workbook['Delays'].cell(row_num, 5, ERC['destination'])
        workbook['Delays'].cell(row_num, 6, ERC['delay'])

    row_num = 1

    for zlt_line in IP['INPUTS']['zlt']:
        row_num += 1
        workbook['Inputs'].cell(row_num, 1, zlt_line)

    row_num = 1

    for zlg_line in IP['INPUTS']['zlg']:
        row_num += 1
        workbook['Inputs'].cell(row_num, 2, zlg_line)

    row_num = 1

    for zop_line in IP['INPUTS']['zop']:
        row_num += 1
        workbook['Inputs'].cell(row_num, 3, zop_line)

    row_num = 1

    for zad_line in IP['INPUTS']['zad']:
        row_num += 1
        workbook['Inputs'].cell(row_num, 4, zad_line)

    if platform.system() == 'Windows':
        save_path = os.getcwd() + '\\stations\\output\\'
    else:
        save_path = os.getcwd() + '/stations/output/'

    file_lbl = IP['COVER']['station_lbl'] + '_Interlocking_Program.xlsx'
    workbook.save(save_path + file_lbl)
